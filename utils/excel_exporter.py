"""
Excel导出工具
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict
from core.bilibili_api import BilibiliAPI


class ExcelExporter:
    """Excel导出类"""
    
    @staticmethod
    def export_ups(ups: Dict[int, Dict], filepath: str) -> bool:
        """
        导出UP主列表到Excel
        
        Args:
            ups: UP主字典
            filepath: 导出路径
        
        Returns:
            bool: 是否成功
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "UP主列表"
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 50
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 50
            
            # 标题行
            headers = ["排名", "UP主名称", "UID", "粉丝数", "投稿数", "等级", "个性签名", "认证", "主页链接"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据行
            sorted_ups = sorted(ups.values(), key=lambda x: x['fans'], reverse=True)
            
            for idx, up in enumerate(sorted_ups, 1):
                row = idx + 1
                
                # 排名
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                
                # UP主名称
                ws.cell(row=row, column=2, value=up['name'])
                ws.cell(row=row, column=2).font = Font(bold=True)
                
                # UID
                ws.cell(row=row, column=3, value=up['mid'])
                
                # 粉丝数
                ws.cell(row=row, column=4, value=up['fans'])
                ws.cell(row=row, column=4).number_format = '#,##0'
                
                # 投稿数
                ws.cell(row=row, column=5, value=up['videos'])
                
                # 等级
                ws.cell(row=row, column=6, value=f"Lv{up['level']}")
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
                
                # 个性签名
                ws.cell(row=row, column=7, value=up['sign'])
                ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
                
                # 认证
                ws.cell(row=row, column=8, value=up['official'] if up['official'] else "无")
                
                # 主页链接（可点击）
                link = f"https://space.bilibili.com/{up['mid']}"
                ws.cell(row=row, column=9, value=link)
                ws.cell(row=row, column=9).hyperlink = link
                ws.cell(row=row, column=9).font = Font(color="0563C1", underline="single")
            
            # 添加统计信息（在数据下方）
            stats_row = len(sorted_ups) + 3
            ws.cell(row=stats_row, column=1, value="导出统计")
            ws.cell(row=stats_row, column=1).font = Font(bold=True, size=11)
            
            ws.cell(row=stats_row + 1, column=1, value="导出时间:")
            ws.cell(row=stats_row + 1, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            ws.cell(row=stats_row + 2, column=1, value="UP主总数:")
            ws.cell(row=stats_row + 2, column=2, value=len(sorted_ups))
            
            # 保存文件
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"[Excel] 导出失败: {e}")
            return False
